import sys
from PyQt5.QtWidgets import QAbstractItemView, QApplication, QButtonGroup, QCompleter, QDialog, QListWidgetItem, QMainWindow, QMessageBox, QTableWidgetItem
from PyQt5.uic import loadUi
from _sqlite3 import Error, IntegrityError, OperationalError
import sqlite3
from PyQt5 import QtCore
from PyQt5.QtGui import QIntValidator
import datetime
from datetime import datetime
from datetime import date
from math import ceil
from openpyxl import Workbook
from PyQt5.QtCore import pyqtSignal
from ventana_buscar_factura_compra import ventana_buscar_factura_compra
from ventana_agenda import ventana_agenda
from openpyxl.reader.excel import load_workbook

class ventana_factura_compra(QMainWindow):
	clicked = pyqtSignal()
	def __init__(self, parent=None):
		super(ventana_factura_compra,self).__init__(parent)
		loadUi('../visual/ventana_carga_facturas.ui', self)

	#Fechas
		self.hoy=QtCore.QDate.currentDate()
		self.fecha_actual.setDate(self.hoy)
		#self.fecha_comp.setDate(hoy)
		#self.fecha_pago.setDate(hoy)

	#listas Combobox
		self.cuenta_imputar.addItems(self.lista_cuenta())
		self.condicion_pago.addItems(self.lista_pagos())
		self.valor_iva.addItems(self.lista_iva())
		
	#Tipo de busqueda
		self.sin_stock.setCheckable(True)
		self.sin_stock.clicked.connect(self.tipo_busqueda)
		self.con_stock.setCheckable(True)
		self.con_stock.clicked.connect(self.tipo_busqueda)

	# Establecer ancho de las columnas
		#Busqueda de productos
		for indice, ancho in enumerate((90, 140, 140, 70), start=0):
			self.tabla_busqueda.setColumnWidth(indice, ancho)

		#Detalle de compra
		for indice, ancho in enumerate((90, 140, 140, 70,40,70,40,40,70,50,70), start=0):
			self.detalle_compra.setColumnWidth(indice, ancho)

	#Configuracion radio botones Stock
		self.grupo_stock= QButtonGroup()
		self.grupo_stock.addButton(self.sin_stock)
		self.grupo_stock.addButton(self.con_stock)       
	
	#señales de widgets
		#Proveedor
		autocompletar = QCompleter(self.ver_proveedor())
		autocompletar.setCaseSensitivity(0)
		self.proveedor.setCompleter(autocompletar)
		self.proveedor.returnPressed.connect(self.datos_proveedor)
		
		#busqueda por producto
		self.nombre_producto.returnPressed.connect(self.busqueda_producto)

		#Señales de botones
		self.cargar_factura.clicked.connect(self.nueva_factura)
		self.buscar_factura.clicked.connect(self.buscar_facturas)
		self.cancelar_carga.clicked.connect(self.cancelar_factura)
		self.nuevo_proveedor.clicked.connect(self.abrir_agenda)
		self.agregar_prod.clicked.connect(self.ingresar)
		self.borrar_fila.clicked.connect(self.borrar)
		self.guardar.clicked.connect(self.guardar_factura)

		#Tabla de busqueda de productos
		self.tabla_busqueda.clicked.connect(self.seleccion_producto)

		#ajuste netos e IVA
		self.ajuste_neto.valueChanged.connect(self.ajusteNeto)
		self.ajuste_iva.valueChanged.connect(self.ajusteIVA)

	#Mostrar última factura hecha	
		self.mostrar_factura(self.ultima_factura())

#Funciones de Botones
	#Nueva carga de factura
	def nueva_factura(self):
		self.limpiar_factura()
		self.anular_factura.setEnabled(False)
		self.cancelar_carga.setEnabled(True)
		self.proveedor.setEnabled(True)
		self.nuevo_proveedor.setEnabled(True)
		self.serie.setEnabled(True)
		self.numero_fact.setEnabled(True)
		self.fecha_comp.setEnabled(True)
		self.fecha_pago.setEnabled(True)
		self.cuenta_imputar.setEnabled(True)
		self.condicion_pago.setEnabled(True)
		self.valor_iva.setEnabled(True)
		self.radio_stock(True)
		self.guardar.setEnabled(True)
		self.fecha_comp.setDate(self.hoy)
		self.fecha_pago.setDate(self.hoy)
		self.id_fact.setText('')
		
	#Activar radiobotones Stock
	def radio_stock(self,value):
		self.sin_stock.setEnabled(value)
		self.con_stock.setEnabled(value)

	#Lista de proveedores
	def ver_proveedor (self):
		conexion=sqlite3.connect('BD.db')
		cursor=conexion.cursor()
		cursor.execute('SELECT nombre FROM proveedor')
		lista=[]
		for i in cursor.fetchall():
			lista.append(i[0])
		return lista

	#Nuevo proveedor	
	def abrir_agenda(self):
		ventana=ventana_agenda(self)
		ventana.show()

	#Tipo de factura segun el proveedor
	def factura_tipo (self):
		if self.tipo_iva.text()=='IVA Responsable Inscripto':
			tipo='"A"'
		elif self.tipo_iva.text()=='Proveedor del Exterior' or self.tipo_iva.text()=='Cliente del Exterior':
			tipo='"E"'
		elif self.tipo_iva.text()=='Consumidor Final':
			QMessageBox.warning(self, 'Error', 'Consumidor final no puede ser proveedor')
		else:
			tipo='"C"'
		return tipo

	#Datos del proveedor
	def datos_proveedor(self):
		nombre=self.proveedor.text()
		conexion=sqlite3.connect('BD.db')
		cursor=conexion.cursor()
		cursor.execute('SELECT * FROM proveedor WHERE nombre=?', [nombre] )
		resultado=cursor.fetchall()
		for i in resultado:
			self.tipo_iva.setText(resultado[0][3])
			self.tipo_doc.setText(resultado[0][4])
			self.numero.setText(resultado[0][5])
			self.direccion.setText(resultado[0][6]+ ' - '+ resultado[0][7])
			self.telefono.setText(resultado[0][8])
			self.email.setText(resultado[0][9])
			self.tipo_fact.setText(self.factura_tipo())
			
	#Lista de tipo de cuenta
	def lista_cuenta (self):
		lista=[
			'',
			'Caja', 'Valores a depositar', 'Valores diferidos a depositar',
			'Tarjeta de crédito', 'Materias primas', 'Muebles y útiles',
			'Banco cuenta corriente', 'Rodados', 'Duedores por venta',
			'Deudores varios', 'Documentos a cobrar', 'Inmuebles',
			'Maquinarias', 'Equipos de computación', 'Documentos a pagar',
			'Acreedores varios', 'Proveedores', 'Valores diferidos a pagar',
			'Capital', 'Comisiones ganadas', 'Gastos generales', 'Costos de mercaderias vendidas',
			'Intereses ganados', 'Venta', 'Descuentos obtenidos', 'Descuentos cedidos',
			'Comisiones perdidas', 'Alquileres ganados', 'Fletes y acarreos', 'Impuestos',
			'Alquileres perdidos', 'Resultado del ejercicio', 'Sueldos y jornales', 'Reservas',
			'Publicidad perdida', 'Seguros'
		]
		return lista
	
	#Lista de tipo de pago
	def lista_pagos (self):
		lista=[
			'',
			'Cuenta Corriente', 'Efectivo', 'Cheques','Tarjeta de Crédito', 
			'Transferencia bancaria'
		]
		return lista

	#Lista de IVA
	def lista_iva (self):
		lista=['','0', '10.5','21','27']
		return lista
		
	#Corroborar si usa articulos del stock o son items sin moviemiento en deposito
	def tipo_busqueda (self):
		if self.sin_stock.isChecked():
			self.desc_prod.setEnabled(True)
			self.desc_prod.setText('')
			self.tabla_busqueda.setEnabled(False)
			self.nombre_producto.setEnabled(False)
			self.nombre_producto.setText('')
			self.tabla_busqueda.clearContents()
			self.marca.setEnabled(False)
			self.marca.setText('')
			self.descripcion.setEnabled(False)
			self.descripcion.setText('')
			self.codigo.setEnabled(False)
			self.codigo.setText('')
			self.cantidad.setEnabled(True)
			self.precio.setEnabled(True)
			self.dto.setEnabled(True)
			self.dto.setText('0')
			self.bonif.setEnabled(True)
			self.bonif.setText('0')
			self.agregar_prod.setEnabled(True)
		elif self.con_stock.isChecked():
			self.tabla_busqueda.setEnabled(True)
			self.nombre_producto.setEnabled(True)
			self.nuevo_producto.setEnabled(True)
			self.desc_prod.setEnabled(False)
			self.desc_prod.setText('')
			self.marca.setEnabled(False)
			self.marca.setText('')
			self.descripcion.setEnabled(False)
			self.descripcion.setText('')
			self.codigo.setEnabled(False)
			self.codigo.setText('')
			self.cantidad.setEnabled(False)
			self.precio.setEnabled(False)
			self.dto.setEnabled(False)
			self.bonif.setEnabled(False)
			self.agregar_prod.setEnabled(False)
		else:
			self.sin_stock.setEnabled(True)
			self.con_stock.setEnabled(True)
		
	#Busqueda productos
	def busqueda_producto (self):
		conexion=sqlite3.connect('BD.db')
		cursor=conexion.cursor()
		dato=self.nombre_producto.text()
		lista=[]
		cursor.execute("SELECT * FROM productos WHERE producto LIKE ? OR marca LIKE ? OR descripcion LIKE ? OR codigo LIKE ?",(f'%{dato}%',f'%{dato}%',f'%{dato}%',f'%{dato}%',))
		datos=cursor.fetchall()
		for i in datos:
			self.tabla_busqueda.clearContents()
		n=0
		for i in datos:
			self.tabla_busqueda.setRowCount(n + 1)
			self.tabla_busqueda.setItem(n, 0, QTableWidgetItem(i[1]))
			self.tabla_busqueda.setItem(n, 1, QTableWidgetItem(i[2]))
			self.tabla_busqueda.setItem(n, 2, QTableWidgetItem(i[3]))
			self.tabla_busqueda.setItem(n, 3, QTableWidgetItem(i[4]))
			n+=1
	
	#Seleccion de producto
	def seleccion_producto (self):
		datos=self.tabla_busqueda.selectedItems()
		marca=datos[0]
		producto=datos[1]
		descripcion=datos[2]
		codigo=datos[3]
		self.marca.setText(marca.text())
		self.desc_prod.setText(producto.text())
		self.descripcion.setText(descripcion.text())
		self.codigo.setText(codigo.text())
		self.cantidad.setEnabled(True)
		self.precio.setEnabled(True)
		self.dto.setEnabled(True)
		self.dto.setText('0')
		self.bonif.setEnabled(True)
		self.bonif.setText('0')
		self.agregar_prod.setEnabled(True)

	#Agregar productos a la tabla final
	def ingresar (self):
		try:
			#Ajustes de widgets
			self.valor_iva.setEnabled(False)
			self.detalle_compra.setEnabled(True)
			self.ajuste_neto.setEnabled(True)
			self.ajuste_neto.setValue(0.00)
			self.ajuste_iva.setEnabled(True)
			self.ajuste_iva.setValue(0.00)
			self.borrar_fila.setEnabled(True)

			#Recoleccion de datos
			cantidad=round(float(self.cantidad.text()),2)
			precio=round(float(self.precio.text()),2)
			if self.dto.text()=='':
				dto=0
			else:
				dto=round(float(self.dto.text()),2)
			if self.bonif.text()=='':
				bonif=0
			else:
				bonif=round(float(self.bonif.text()),2)
			valor_dto=precio*(dto*0.01)
			valor_bonif=precio*(bonif*0.01)
			importe=round(cantidad*(precio-valor_bonif-valor_dto),2)
			iva=round(((float(self.valor_iva.currentText())+100)/100),2)
			total=round(importe*iva,2)

			#Empaquetado de datos para tabla
			datos=[(self.marca.text(),
				self.desc_prod.text(),
				self.descripcion.text(),
				self.codigo.text(),
				self.cantidad.text(),
				self.precio.text(),
				self.dto.text(),
				self.bonif.text(),
				str(importe),
				self.valor_iva.currentText(),
				str(total)
			)]
			
			#Armado de tabla
			fila= self.detalle_compra.rowCount()
			n=fila
			for i in datos:
				self.detalle_compra.setRowCount(n+1)
				self.detalle_compra.setItem(n, 0, QTableWidgetItem(i[0]))
				self.detalle_compra.setItem(n, 1, QTableWidgetItem(i[1]))
				self.detalle_compra.setItem(n, 2, QTableWidgetItem(i[2]))
				self.detalle_compra.setItem(n, 3, QTableWidgetItem(i[3]))
				self.detalle_compra.setItem(n, 4, QTableWidgetItem(i[4]))
				self.detalle_compra.setItem(n, 5, QTableWidgetItem(i[5]))
				self.detalle_compra.setItem(n, 6, QTableWidgetItem(i[6]))
				self.detalle_compra.setItem(n, 7, QTableWidgetItem(i[7]))
				self.detalle_compra.setItem(n, 8, QTableWidgetItem(i[8]))
				self.detalle_compra.setItem(n, 9, QTableWidgetItem(i[9]))
				self.detalle_compra.setItem(n, 10, QTableWidgetItem(i[10]))
				n+=1

			#Limpieza para volver a empezar	
			self.marca.setText(''),
			self.desc_prod.setText('')
			self.descripcion.setText('')
			self.codigo.setText('')
			self.cantidad.setText('')
			self.precio.setText('')
			self.dto.setText('0')
			self.bonif.setText('0')

		except ValueError:
			QMessageBox.warning(self, 'Error', 'Verifique que los campos "Cantidad" y "Precio", se encuentren completados')

		#Suma de montos
		self.suma_neto()
		self.suma_iva()
		self.suma_total()
		
	#Ajuste stock
	def ajuste_stock (self):
		lista=[]
		filas=self.detalle_compra.rowCount()
		for i in range(filas):
			ingreso=[]
			ingreso.append(self.detalle_compra.item(i, 3).text())
			ingreso.append(self.detalle_compra.item(i, 4).text())
			ingreso.append(self.detalle_compra.item(i,5).text())
			stock=(ingreso[0],ingreso[1],ingreso[2])
			lista.append(stock)
		return lista

	#Suma de Neto
	def suma_neto (self):
		item = []
		filas=self.detalle_compra.rowCount()
		for i in range(filas):
			item.append(float(self.detalle_compra.item(i, 8).text()))
		suma=round(sum(item),2)
		neto=self.neto.setText(str(suma))

	#Suma de IVA
	def suma_iva (self):
		neto=float(self.neto.text())
		iva=round(neto*(float(self.valor_iva.currentText())*0.01),2)
		valor_iva=self.iva.setText(str(iva))

	#Suma de Total
	def suma_total (self):
		neto=float(self.neto.text())
		iva=float(self.iva.text())
		total=round((neto+iva),2)
		self.total.setText(str(total))
	
	#Suma descuento y bonificación
	def suma_dto(self):
		dto = []
		bonif=[]
		filas=self.detalle_compra.rowCount()
		for i in range(filas):
			dto.append((float(self.detalle_compra.item(i, 4).text()))*(float(self.detalle_compra.item(i, 5).text())*(float(self.detalle_compra.item(i, 6).text())*0.01)))
			bonif.append((float(self.detalle_compra.item(i, 4).text()))*(float(self.detalle_compra.item(i, 5).text())*(float(self.detalle_compra.item(i, 7).text())*0.01)))
		suma_dto=round(sum(dto),2)
		suma_bonif=round(sum(bonif),2)
		lista=[suma_dto,suma_bonif]
		return lista

	#Ajustar decimales en Neto	
	def ajusteNeto (self):
		self.suma_neto()
		ajuste=self.ajuste_neto.value()
		neto=float(self.neto.text())
		nuevo_neto=neto+ajuste
		self.neto.setText(str(nuevo_neto))
		self.suma_total()

	#Ajustar decimales en IVA	
	def ajusteIVA (self):
		self.suma_iva()
		ajuste=self.ajuste_iva.value()
		iva=float(self.iva.text())
		nuevo_iva=round((iva+ajuste),2)
		self.iva.setText(str(nuevo_iva))
		self.suma_total()

	#Borrar fila seleccionada
	def borrar (self):
		seleccion=self.detalle_compra.selectedItems()	
		fila=self.detalle_compra.row(seleccion[0])
		self.detalle_compra.removeRow(fila)
		self.suma_neto()
		self.suma_iva()
		self.suma_total()

	#Guardar
	def guardar_factura (self):
		conexion=sqlite3.connect('BD.db')
		cursor=conexion.cursor()
		numero=self.serie.text()+'-'+self.numero_fact.text()
		fecha_comp=("%s/%s/%s"%(self.fecha_comp.date().day(),
								   self.fecha_comp.date().month(),
								   self.fecha_comp.date().year()))
		fecha_pago=("%s/%s/%s"%(self.fecha_pago.date().day(),
								   self.fecha_pago.date().month(),
								   self.fecha_pago.date().year()))
		
		#Calculos de descuentos bonificaciones
		descuento=self.suma_dto()[0]
		bonificacion=self.suma_dto()[1]
		
		if self.sin_stock.isChecked()==True:
			datos=[
				self.proveedor.text(),
				self.tipo_fact.text(),
				numero,
				fecha_comp,
				fecha_pago,
				self.cuenta_imputar.currentText(),
				self.condicion_pago.currentText(),
				self.neto.text(),
				self.iva.text(),
				self.total.text(),
				descuento,
				bonificacion,
				'No'
			]
			cursor.execute('INSERT INTO facturas_compra VALUES(NULL,?,?,?,?,?,?,?,?,?,?,?,?,?)', (datos))
			conexion.commit()
		elif self.con_stock.isChecked()==True:
			datos=[
				self.proveedor.text(),
				self.tipo_fact.text(),
				numero,
				fecha_comp,
				fecha_pago,
				self.cuenta_imputar.currentText(),
				self.condicion_pago.currentText(),
				self.neto.text(),
				self.iva.text(),
				self.total.text(),
				descuento,
				bonificacion,
				'Si'
			]
			cursor.execute('INSERT INTO facturas_compra VALUES(NULL,?,?,?,?,?,?,?,?,?,?,?,?,?)', (datos))
			conexion.commit()
			stock=self.ajuste_stock()
			for i in stock:
				cursor.execute('SELECT cantidad FROM productos WHERE codigo=?', [i[0]])
				info=cursor.fetchall()
				datos=[]
				datos.append(info[0][0])
				datos.append(float(i[1]))
				ajuste=sum(datos)
				precio_nuevo=float(i[2])
				print(ajuste,precio_nuevo,i[0])
				cursor.execute('UPDATE productos SET cantidad=?, precio=? WHERE codigo=?', (ajuste,precio_nuevo,i[0]))
				conexion.commit()
			QMessageBox.information(self, 'Importante','Se modificaron costos, por favor corregir precio de ventas', QMessageBox.Ok, QMessageBox.Ok)
				
		else:
			QMessageBox.warning(self,'Error', 'Algo no funciona correctamente si ves esto')
		self.datos_factura_compra()
		self.inicio_fact()

	#datos de factura en excel
	def datos_factura_compra(self):
		nro=f'{self.serie.text()}-{self.numero_fact.text()}'
		nombre=str(f'N° {nro} - {self.proveedor.text()}')
		archivo = Workbook()
		hoja=archivo.active
		hoja['A2']='Factura N°: '
		hoja['A3']=self.serie.text()
		hoja['B3']=self.numero_fact.text()
		hoja['C2']='Fecha:'
		hoja['C3']=self.fecha_comp.text()
		hoja['D2']='Proveedor:'
		hoja['D3']=self.proveedor.text()
		hoja['B5']='Neto: '
		hoja['B6']=self.neto.text()
		hoja['C5']= 'IVA:'
		hoja['C6']=	self.iva.text()
		hoja['D5']='Total:'
		hoja['D6']=self.total.text()
		hoja['B7']=''
		hoja.append(['marca','producto','descripcion', 'codigo', 'cant.', 'precio', '"%" dto.','"%" bonif.', 'importe','iva', 'total'])
		tabla = []
		fila = self.detalle_compra.rowCount()
		for f in range(fila):
			datos=[]
			for c in range(11):
				datos.append(self.detalle_compra.item(f, c).text())
			tabla.append(datos)
		for i in tabla:
			hoja.append(i)

		#Ajuste de ancho de columnas
		for i in hoja.columns:
			ancho = 0
			columna = i[0].column_letter
			for celda in i:
				try:
					if len(str(celda.value)) > ancho:
						ancho = len(celda.value)
				except:
					pass
			ajuste_celda = (ancho + 2) * 1.2
			hoja.column_dimensions[columna].width = ajuste_celda

		#Guardado del archivo
		archivo.save(f"../archivos/factura_compra/{nombre}.xlsx")

	#Volver al estado inicial
	def inicio_fact (self):
		self.proveedor.setEnabled(False)
		self.nuevo_proveedor.setEnabled(False)
		self.serie.setEnabled(False)
		self.numero_fact.setEnabled(False)
		self.fecha_comp.setEnabled(False)
		self.radio_stock(False)
		self.nombre_producto.setEnabled(False)
		self.nuevo_producto.setEnabled(False)
		self.tabla_busqueda.setEnabled(False)
		self.desc_prod.setEnabled(False)
		self.cantidad.setEnabled(False)
		self.precio.setEnabled(False)
		self.dto.setEnabled(False)
		self.bonif.setEnabled(False)
		self.agregar_prod.setEnabled(False)
		self.borrar_fila.setEnabled(False)
		self.detalle_compra.setEnabled(False)
		self.ajuste_neto.setEnabled(False)
		self.ajuste_iva.setEnabled(False)
		self.guardar.setEnabled(False)
		self.fecha_pago.setEnabled(False)
		self.cuenta_imputar.setEnabled(False)
		self.condicion_pago.setEnabled(False)
		self.valor_iva.setEnabled(False)
		self.guardar.setEnabled(False)

	#Cancelar la carga de la factura
	def cancelar_factura(self):
		self.limpiar_factura()
		self.mostrar_factura(self.ultima_factura())
	
	#Eliminar filas de tabla contactos
	def limpiar_filas(self):
		filas=self.tabla_busqueda.rowCount()
		for i in range(filas):
			self.tabla_busqueda.removeRow(0)
		self.tabla_busqueda.clearContents()
		filas=self.detalle_compra.rowCount()
		for i in range(filas):
			self.detalle_compra.removeRow(0)
		self.detalle_compra.clearContents()

	#Limpiar pantalla
	def limpiar_factura (self):
		self.proveedor.setText('')
		self.tipo_iva.setText('')
		self.tipo_doc.setText('')
		self.numero.setText('')
		self.direccion.setText('')
		self.telefono.setText('')
		self.email.setText('')
		self.tipo_fact.setText('')
		self.serie.setText('')
		self.numero_fact.setText('')
		self.cuenta_imputar.setCurrentText('')
		self.condicion_pago.setCurrentText('')
		self.valor_iva.setCurrentText('')
		self.nombre_producto.setText('')
		self.desc_prod.setText('')
		self.marca.setText('')
		self.descripcion.setText('')
		self.codigo.setText('')
		self.cantidad.setText('')
		self.precio.setText('')
		self.dto.setText('')
		self.bonif.setText('')
		self.neto.setText('')
		self.iva.setText('')
		self.total.setText('')
		self.grupo_stock.setExclusive(False)   
		self.sin_stock.setChecked( False )
		self.con_stock.setChecked( False )
		self.grupo_stock.setExclusive(True)   
		self.tabla_busqueda.clearContents()
		self.detalle_compra.clearContents()
		self.limpiar_filas()
		self.anular_factura.setEnabled(True)
		self.cancelar_carga.setEnabled(False)
		self.inicio_fact()

	#Buscar facturas existentes
	def buscar_facturas(self):
		ventana=ventana_buscar_factura_compra(self)
		ventana.datos_factura.connect(self.mostrar_factura)
		ventana.show()

	#Cargar datos de la busqueda de datos
	def mostrar_factura(self,dateStr):
		datos_busqueda=dateStr
		id_factura=datos_busqueda[0]
		conexion=sqlite3.connect('BD.db')
		cursor=conexion.cursor()
		cursor.execute('SELECT * FROM facturas_compra WHERE id=?',id_factura)
		datos=cursor.fetchall()
		conexion.close()
		#Distribucion de datos en la pantalla
		self.proveedor.setText(datos[0][1])
		self.datos_proveedor()
		self.id_fact.setText(id_factura)
		#Numeros de factura
		numero=str(datos[0][3]).split(sep='-')
		self.serie.setText(numero[0])
		self.numero_fact.setText(numero[1])
		#Fecha de carga
		fecha_carga=datos[0][4]
		cambio_fecha=datetime.strptime(fecha_carga,'%d/%m/%Y')
		self.fecha_comp.setDate(cambio_fecha)
		#fecha de pago
		fecha_pago=datos[0][5]
		cambio_fecha=datetime.strptime(fecha_pago,'%d/%m/%Y')
		self.fecha_pago.setDate(cambio_fecha)
		# cuenta a imputar
		cuenta=datos[0][6]
		self.cuenta_imputar.setCurrentText(cuenta)
		#condicion de pago
		pago=datos[0][7]
		self.condicion_pago.setCurrentText(pago)
		
		#Recuperación del detalle de excel
		nro=f'{self.serie.text()}-{self.numero_fact.text()}'
		nombre=str(f'N° {nro} - {self.proveedor.text()}')
		archivo = load_workbook(f"../archivos/factura_compra/{nombre}.xlsx")
		hoja=archivo.active
		datos_factura=[]
		for i in hoja.iter_rows(min_row=9,min_col=1,max_col=11, values_only=True):
			datos_factura.append(i)
			#Armado de tabla
		fila= self.detalle_compra.rowCount()
		n=fila
		for i in datos_factura:
			self.detalle_compra.setRowCount(n+1)
			self.detalle_compra.setItem(n, 0, QTableWidgetItem(i[0]))
			self.detalle_compra.setItem(n, 1, QTableWidgetItem(i[1]))
			self.detalle_compra.setItem(n, 2, QTableWidgetItem(i[2]))
			self.detalle_compra.setItem(n, 3, QTableWidgetItem(i[3]))
			self.detalle_compra.setItem(n, 4, QTableWidgetItem(i[4]))
			self.detalle_compra.setItem(n, 5, QTableWidgetItem(i[5]))
			self.detalle_compra.setItem(n, 6, QTableWidgetItem(i[6]))
			self.detalle_compra.setItem(n, 7, QTableWidgetItem(i[7]))
			self.detalle_compra.setItem(n, 8, QTableWidgetItem(i[8]))
			self.detalle_compra.setItem(n, 9, QTableWidgetItem(i[9]))
			self.detalle_compra.setItem(n, 10, QTableWidgetItem(i[10]))
			n+=1
		
		#Suma de montos
		#self.suma_neto()
		#self.suma_iva()
		#self.suma_total()

		'''
		self.tablaLista.delete(*self.tablaLista.get_children())
		self.seleccion=self.tablaCte.item(self.tablaCte.selection())['values']
		titulo=f'{self.seleccion[1]}-{self.seleccion[3]}'
		archivo = load_workbook(f"./Comprobantes/{titulo}.xlsx")
		hoja=archivo.active
		art=[]
		for i in hoja.iter_rows(min_row=10,min_col=4,max_col=4, values_only=True):
			art.append(i[0])
			
		cant=[]
		for e in hoja.iter_rows(min_row=10,min_col=5,max_col=5, values_only=True):
			cant.append(e[0])
				
		lista=dict(zip(art,cant))
		elementos=lista.items()
		for (i,e) in elementos:
			self.tablaLista.insert('', 'end', values=(i,e))'''

	#Retornar el ID del último comprobante para mostrarlo cuando no haya carga
	def ultima_factura(self):
		conexion=sqlite3.connect('BD.db')
		cursor=conexion.cursor()
		cursor.execute('SELECT id FROM facturas_compra ORDER by id DESC')
		ultima=cursor.fetchone()
		conexion.close()
		return str(ultima[0])
		

#Cierre de la app para que se ejecute
app = QApplication(sys.argv)
app.setStyle('Fusion')
main = ventana_factura_compra()
main.show()
sys.exit(app.exec_())