import sys
from PyQt5.QtWidgets import QAbstractItemView, QApplication, QButtonGroup, QCompleter, QDialog, QListWidgetItem, QMainWindow, QMessageBox, QTableWidgetItem
from PyQt5.uic import loadUi
from _sqlite3 import Error, IntegrityError, OperationalError
import sqlite3
from PyQt5 import QtCore
from PyQt5.QtGui import QIntValidator
import datetime
from datetime import datetime
from datetime import date
from math import ceil
from openpyxl import Workbook
from PyQt5.QtCore import pyqtSignal
from openpyxl.reader.excel import load_workbook
import time


class ventana_facturas_anuladas(QMainWindow):
	clicked = pyqtSignal()
	def __init__(self, parent=None):
		super(ventana_facturas_anuladas,self).__init__(parent)
		loadUi('visual/ventana_facturas_anuladas.ui', self)

		#Fechas
		hoy=QtCore.QDate.currentDate()
		self.fecha_comp.setDate(hoy)
		self.fecha_pago.setDate(hoy)

		#busqueda de facturas
		self.busqueda_anulada.returnPressed.connect(self.busqueda_facturas)

		# Establecer ancho de las columnas
		#Busqueda de factura anulada
		for indice, ancho in enumerate((50, 150, 120, 100), start=0):
			self.tabla_busqueda.setColumnWidth(indice, ancho)

		#Detalle de factura
		for indice, ancho in enumerate((40,90, 140, 180, 70,40,70,40,50,70,40,70), start=0):
			self.detalle_compra.setColumnWidth(indice, ancho)

		#Tabla de busqueda de facturas
		self.tabla_busqueda.clicked.connect(self.mostrar_detalle)

#Funciones de Botones
	#Busqueda productos
	def busqueda_facturas(self):
		self.limpiar_busqueda()
		self.tabla_busqueda.clearContents()
		conexion=sqlite3.connect('BD.db')
		cursor=conexion.cursor()
		dato=self.busqueda_anulada.text()
		lista=[]
		cursor.execute("SELECT * FROM facturas_anuladas WHERE id LIKE ? OR proveedor LIKE ? OR numero LIKE ? OR total LIKE ?",(f'%{dato}%',f'%{dato}%',f'%{dato}%',f'%{dato}%',))
		datos=cursor.fetchall()
		for i in datos:
			self.tabla_busqueda.clearContents()
		n=0
		for i in datos:
			self.tabla_busqueda.setRowCount(n + 1)
			self.tabla_busqueda.setItem(n, 0, QTableWidgetItem(str(i[0])))
			self.tabla_busqueda.setItem(n, 1, QTableWidgetItem(i[1]))
			self.tabla_busqueda.setItem(n, 2, QTableWidgetItem(i[3]))
			self.tabla_busqueda.setItem(n, 3, QTableWidgetItem(i[10]))
			n+=1

	#Abrir factura seleccionada
	def mostrar_detalle(self):
		seleccion=self.tabla_busqueda.selectedItems()
		if seleccion!=[]:
		#Cargar datos de la busqueda de datos
			datos_busqueda=seleccion[0].text()
			id_factura=datos_busqueda
			conexion=sqlite3.connect('BD.db')
			cursor=conexion.cursor()
			cursor.execute('SELECT * FROM facturas_anuladas WHERE id=?',[id_factura])
			datos=cursor.fetchall()
			conexion.close()
			#Distribucion de datos en la pantalla
			self.proveedor.setText(datos[0][1])

			#Numeros de factura
			numero=str(datos[0][3]).split(sep='-')
			self.serie.setText(numero[0])
			self.numero_fact.setText(numero[1])
			#Fecha de carga
			fecha_carga=datos[0][4]
			
			cambio_fecha=datetime.strptime(fecha_carga,'%d/%m/%Y')
			self.fecha_comp.setDate(cambio_fecha)
			#fecha de pago
			fecha_pago=datos[0][5]
			cambio_fecha=datetime.strptime(fecha_pago,'%d/%m/%Y')
			self.fecha_pago.setDate(cambio_fecha)
			# cuenta a imputar
			cuenta=datos[0][6]
			self.cuenta_imputar.setText(cuenta)
			#condicion de pago
			pago=datos[0][7]
			self.condicion_pago.setText(pago)

			#Observaciones de anulacion
			obs=datos[0][15]
			self.detalle_anulacion.setPlainText(obs)

			#Recuperación del detalle de excel
			nro=f'{self.serie.text()}-{self.numero_fact.text()}'
			nombre=str(f'N° {nro} - {self.proveedor.text()}')
			archivo = load_workbook(f"archivos/factura_compra/{nombre}.xlsx")
			hoja=archivo.active
			datos_factura=[]
			for i in hoja.iter_rows(min_row=9,min_col=1,max_col=12, values_only=True):
				datos_factura.append(i)
				#Armado de tabla
			self.limpiar_detalle()
			fila= self.detalle_compra.rowCount()
			n=fila
			for i in datos_factura:
				self.detalle_compra.setRowCount(n+1)
				self.detalle_compra.setItem(n, 0, QTableWidgetItem(i[0]))
				self.detalle_compra.setItem(n, 1, QTableWidgetItem(i[1]))
				self.detalle_compra.setItem(n, 2, QTableWidgetItem(i[2]))
				self.detalle_compra.setItem(n, 3, QTableWidgetItem(i[3]))
				self.detalle_compra.setItem(n, 4, QTableWidgetItem(i[4]))
				self.detalle_compra.setItem(n, 5, QTableWidgetItem(i[5]))
				self.detalle_compra.setItem(n, 6, QTableWidgetItem(i[6]))
				self.detalle_compra.setItem(n, 7, QTableWidgetItem(i[7]))
				self.detalle_compra.setItem(n, 8, QTableWidgetItem(i[8]))
				self.detalle_compra.setItem(n, 9, QTableWidgetItem(i[9]))
				self.detalle_compra.setItem(n, 10, QTableWidgetItem(i[10]))
				self.detalle_compra.setItem(n, 11, QTableWidgetItem(i[11]))
				n+=1

			#Tipo de factura
			tipo_fact=[]
			for i in hoja.iter_rows(min_row=2,max_row=2,min_col=2,max_col=2, values_only=True):
				tipo_fact.append(i)
			cambio=(tipo_fact[0][0]).split(sep='Tipo: ')
			self.tipo_fact.setText(cambio[1])

			#Neto
			neto=[]
			for i in hoja.iter_rows(min_row=6,max_row=6,min_col=2,max_col=2, values_only=True):
				neto.append(i)
			self.neto.setText(neto[0][0])

			#IVA
			iva=[]
			for i in hoja.iter_rows(min_row=6,max_row=6,min_col=34,max_col=34, values_only=True):
				iva.append(i)
			self.iva.setText(iva[0][0])

			#Total
			total=[]
			for i in hoja.iter_rows(min_row=6,max_row=6,min_col=4,max_col=4, values_only=True):
				total.append(i)
			self.total.setText(total[0][0])
			
		else:
			QMessageBox.warning(self, 'Error', 'No se ha seleccionado ningún comprobante')

	#Eliminar filas de la busqueda
	def limpiar_busqueda(self):
		self.tabla_busqueda.clearContents()	
		filas=self.tabla_busqueda.rowCount()
		for i in range(filas):
			self.tabla_busqueda.removeRow(0)

	#Eliminar filas del detalle
	def limpiar_detalle(self):
		self.detalle_compra.clearContents()	
		filas=self.detalle_compra.rowCount()
		for i in range(filas):
			self.detalle_compra.removeRow(0)
'''
#Cierre de la app para que se ejecute
app = QApplication(sys.argv)
app.setStyle('Fusion')
main = ventana_facturas_anuladas()
main.show()
sys.exit(app.exec_())'''