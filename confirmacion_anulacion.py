import sys
from PyQt5.QtWidgets import QAbstractItemView, QApplication, QButtonGroup, QCompleter, QDialog, QDialogButtonBox, QListWidgetItem, QMainWindow, QMessageBox, QTableWidgetItem
from PyQt5.uic import loadUi
from _sqlite3 import Error, IntegrityError, OperationalError
import sqlite3
from PyQt5 import QtCore
from PyQt5.QtGui import QIntValidator
import datetime
from datetime import datetime
from datetime import date
from math import ceil
from openpyxl import Workbook
from PyQt5.QtCore import pyqtSignal
from openpyxl.reader.excel import load_workbook
import time

class confirmacion_anulacion(QDialog):
	clicked = pyqtSignal()
	def __init__(self, parent=None):
		super(confirmacion_anulacion,self).__init__(parent)
		loadUi('visual/confirmacion_anulacion.ui', self)

		#Traducir texto de los botones
		cancelar = self.buttonBox.button(QDialogButtonBox.Cancel)
		cancelar.setText("Cancelar")
		cancelar.clicked.connect(self.cerrar)

		aceptar = self.buttonBox.button(QDialogButtonBox.Ok)
		aceptar.setText("Aceptar")
		aceptar.clicked.connect(self.confirmar_anulacion)

	def cerrar(self):
		pass

	def confirmar_anulacion(self):
		id_fact=self.id_fact.text()
		observacion=self.observacion.toPlainText()

		#Fecha de anulación de factura
		fecha_anulacion=(time.strftime("%d/%m/%y %H:%M:%S"))

		conexion=sqlite3.connect('BD.db')
		cursor=conexion.cursor()
		cursor.execute('SELECT * FROM facturas_compra WHERE id=?', [id_fact])
		datos=list(cursor.fetchall()[0])
		datos.append(observacion)
		datos.append(fecha_anulacion)
		datos.pop(0)

		cursor.execute('INSERT INTO facturas_anuladas VALUES(NULL,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',datos)
		
		stock=self.ajuste_stock()
		for i in stock:
			if datos[14]=='Si':
				cursor.execute('SELECT cantidad FROM productos WHERE id=? ', [i[0]])
				info=cursor.fetchall()
				datos=[]
				datos.append(info[0][0])
				datos.append(-(float(i[5])))
				ajuste=sum(datos)
				cursor.execute('UPDATE productos SET cantidad=? WHERE id=?', (ajuste,i[0]))
		else:
			pass
			
		cursor.execute('DELETE FROM facturas_compra WHERE id=?', [id_fact])
		conexion.commit()
		conexion.close()
		
		QMessageBox.information(self, 'Correcto','Se anuló satisfactoriamente la factura', QMessageBox.Ok, QMessageBox.Ok) #El llamado de self es sin comillas
		
	
	#Ajuste stock
	def ajuste_stock (self):
		lista=[]
		#Recuperación del detalle de excel
		nro=f'{self.serie.text()}-{self.numero_fact.text()}'
		nombre=str(f'N° {nro} - {self.proveedor.text()}')
		archivo = load_workbook(f"archivos/factura_compra/{nombre}.xlsx")
		hoja=archivo.active
		detalle_factura=[]
		for i in hoja.iter_rows(min_row=9,min_col=1,max_col=12, values_only=True):
			detalle_factura.append(i)
		
		items=len(detalle_factura)
		for i in range(items):
			lista.append(detalle_factura[i])
						
		return lista

'''
#Cierre de la app para que se ejecute
app = QApplication(sys.argv)
app.setStyle('Fusion')
main = confirmacion_anulacion()
main.show()
sys.exit(app.exec_())'''